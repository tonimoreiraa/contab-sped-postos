import openpyxl
from save_to_sheet import save_to_sheet

cnpj = "31391653000172"
empresa = "COMERCIAL VERDES CAMPOS LTDA"

def get_data():
    try:
        file_path = f"input/relatorio/{cnpj}.pdf"
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except:
        file_path = f"input/relatorio/{cnpj}.xlsx"
        wb = openpyxl.load_workbook(file_path, data_only=True)
         
    sheet = wb.active

    init = None
    end = None

    for row in sheet.iter_rows():
        if "BOMBA" in str(row[0].value):
            init = row[0].row
        elif "TANTQUE" in str(row[0].value):
            end = row[0].row
            break  

    list_of_dicts = []

    if init is not None and end is not None and init < end:
        for i in range(init + 1, end):  # init + 1 to skip the header
            row_dict = {cell.column: cell.value for cell in sheet[i] if cell.value is not None}
            list_of_dicts.append(row_dict)  
    else:
        print("Não foi possível encontrar as linhas inicial ou final, ou o intervalo é inválido.")

    bico_data = []
    for list in list_of_dicts:
        if list:
            try:
                bico_data.append({
                    'Bico':list[4],
                    'Produto':list[5],
                    'Abertura':list[6],
                    'Fechamento':list[7],
                    'Sem_intervencao':None,
                    'Com_intervencao':None,
                    #'Lacre':list[3],
                    'Afericao':list[9]
                })
            except:
                pass
    
    list_of_dicts = []

    for i in range(end + 1, sheet.max_row):  # end + 1 to skip the header
        row_dict = {cell.column: cell.value for cell in sheet[i] if cell.value is not None}
        list_of_dicts.append(row_dict)

    tanque_data = []
    for list in list_of_dicts:
        if list:
            try:
                tanque_data.append({
                    'Tanque':list[1],
                    'Produto':list[2],
                    'Abertura':list[5],
                    'Fechamento':list[7],
                    'Recebimento':list[6]
                })
            except:
                pass

    save_to_sheet(bico_data, tanque_data, f"output/{cnpj}.xlsx")
    path_dac = f"input/dac/{cnpj}.txt"
    return bico_data, tanque_data, empresa, path_dac