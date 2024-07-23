import openpyxl

def get_data():
    arquivo = 'ap acqua bool.xlsx'
    wb = openpyxl.load_workbook(arquivo, data_only=True)  
    sheet = wb.active

    init = None
    end = None

    for row in sheet.iter_rows():
        if "MOVIMENTAÇÃO POR BICO DE COMBUSTIVEL" in str(row[0].value):
            init = row[0].row
        elif "ESTOQUE" in str(row[0].value):
            end = row[0].row
            break  

    list_of_dicts = []

    if init is not None and end is not None and init < end:
        for i in range(init + 2, end):  # init + 2 to skip the header
            row_dict = {cell.column: cell.value for cell in sheet[i] if cell.value is not None}
            list_of_dicts.append(row_dict)  
    else:
        print("Não foi possível encontrar as linhas inicial ou final, ou o intervalo é inválido.")

    bicos = []
    for list in list_of_dicts:
        if list:
            bicos.append({
                'Serie':list[1],
                'Bico':list[2],
                'Produto':list[3],
                'Abertura':list[5],
                'Fechamento':list[6],
                'Sem_intervencao':list[8],
                'Com_intervencao':list[9],
                'Lacre':list[10],
                'Afericao':list[11]
            })

    list_of_dicts = []

    for i in range(end + 2, sheet.max_row):  # end + 2 to skip the header
        row_dict = {cell.column: cell.value for cell in sheet[i] if cell.value is not None}
        list_of_dicts.append(row_dict)

    tanques = []
    for list in list_of_dicts:
        if list:
            tanques.append({
                'Tanque':list[1],
                'Produto':list[2],
                'Abertura':list[5],
                'Fechamento':list[8],
                'Recebimento':list[10]
            })

    return bicos, tanques