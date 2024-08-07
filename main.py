import os
from input import read_input
from daily_check import daily_check
from identify_models import identify_report_model
from models import models
from comparate_reports import comparate
from openpyxl import Workbook

input_path = 'input/dac'
report_path = 'input/relatorio'

files_and_dirs = os.listdir(input_path)
cnpjs = [f.replace('.txt', '') for f in files_and_dirs if os.path.isfile(os.path.join(input_path, f))]

for cnpj in cnpjs:
    input_data = read_input(os.path.join(input_path, cnpj + '.txt'))
    if input_data:
        daily_output = daily_check(input_data)

        report_file_path = os.path.join(report_path, cnpj)
        if os.path.exists(report_file_path + '.pdf'):
            report_file_path = report_file_path + '.pdf'
        else:
            report_file_path = report_file_path + '.xlsx'

        report_model = identify_report_model(report_file_path)

        report_output = None
        if report_model != 'Modelo desconhecido':
            print(report_model, cnpj)
            report_data = models[report_model](cnpj, report_file_path)
            report_output = comparate(input_data, report_data)
        else:
            print(f'Relatório do CNPJ {cnpj} não foi reconhecido como um modelo cadastrado')

        # save output
        output_path = 'output/' + cnpj + '.xlsx'
        wb = Workbook()
        wb.remove(wb.active)
        if daily_output and len(daily_output):
            ws = wb.create_sheet(title='Conferência diária')
            ws.append(list(daily_output[0].keys()))
            for entry in daily_output:
                ws.append(list(entry.values()))

        if report_output:
            ws = wb.create_sheet(title='Conferência geral')
            ws.append(list(report_output[0].keys()))
            for entry in report_output:
                ws.append(list(entry.values()))
                
        if daily_output or report_output:
            wb.save(output_path)