import openpyxl
from format_value import format_value

def get_data(cnpj, file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)

    sheet = wb.active
    sheet = wb["Plan1"]

    init = None
    end = None
    for row in sheet.iter_rows():
        if "MOVIMENTAÇÃO DE BOMBAS BICOS E LACRES" in str(row[0].value):
            init = row[0].row
        elif "INFORMAÇÕES MENSAIS DE ESTOQUES DE COMBUSTÍVEIS" in str(row[0].value):
            end = row[0].row
            break  

    list_of_dicts = []

    if init is not None and end is not None and init < end:
        for i in range(init + 10, end):  # init + 8 to skip the header
            row_dict = {cell.column: cell.value for cell in sheet[i] if cell.value is not None}
            list_of_dicts.append(row_dict)  
    else:
        print("Não foi possível encontrar as linhas inicial ou final, ou o intervalo é inválido.")

    bico_data = []
    for list in list_of_dicts:
        if list:
            try:
                bico_data.append({
                    'type':'bico',
                    'bico':list[5],
                    #'produto':list[6],
                    'abertura':format_value(list[8]),
                    'fechamento':format_value(list[9]),
                    #'Sem_intervencao':list[10],
                    #'Com_intervencao': None,
                    #'Lacre':list[12],
                    'afericao':format_value(list[13])
                })
            except:
                pass

    list_of_dicts = []

    for i in range(end + 9, sheet.max_row):  # end + 8 to skip the header
        row_dict = {cell.column: cell.value for cell in sheet[i] if cell.value is not None}
        list_of_dicts.append(row_dict)

    tanque_data = []
    for list in list_of_dicts:
        if list:
            try:
                tanque_data.append({
                    'type':'tanque',
                    'tanque':list[1],
                    #'Produto':list[2],
                    'abertura':format_value(list[4]),
                    'fechamento':format_value(list[7]),
                    'recebimento':format_value(list[9])
                })
            except:
                pass

    bico_tanque_data = []
    for bico in bico_data:
        bico_tanque_data.append(bico)
    for tanque in tanque_data:
        bico_tanque_data.append(tanque)

    path_dac = f"input/dac/{cnpj}.txt"
    path_xlsx = f"output/{cnpj}.xlsx"
    return bico_tanque_data, cnpj, path_dac, path_xlsx