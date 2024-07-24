from openpyxl import load_workbook

wb = load_workbook('input/relatorio/10800554000127.xlsx', data_only=True)

sheet = wb.active
sheet = wb["Plan1"]

for row in sheet.iter_rows():
    if row[6].value == 344:
        row[6].value = 200

wb.save('input/relatorio/10800554000127.xlsx')