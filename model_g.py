import openpyxl
from format_value import format_value

def get_data(cnpj, file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
      
    sheet = wb.active

    init = None
    end = None

    for row in sheet.iter_rows():
        if "MOVIMENTAÇÃO POR BICO DE COMBUSTIVEL" in str(row[0].value):
            init = row[0].row
        elif "ESTOQUE" in str(row[0].value):
            end = row[0].row
            break
        elif "]" in str(row[0].value):
            end = row[0].row
            break

    list_of_dicts = []

    if init is not None and end is not None and init < end:
        for i in range(init + 2, end):  # init + 2 to skip the header
            row_dict = {cell.column: cell.value for cell in sheet[i] if cell.value is not None}
            list_of_dicts.append(row_dict)  
    else:
        print("Não foi possível encontrar as linhas inicial ou final, ou o intervalo é inválido.")

    bico_data = []
    for list in list_of_dicts:
        if list:
            try:
                bico_data.append({
                    #'Serie':list[1],
                    'type':'bico',
                    'bico':list[2],
                    #'Produto':list[3],
                    'abertura':format_value(list[5]),
                    'fechamento':format_value(list[6]),
                    #'Sem_intervencao':list[8],
                    #'Com_intervencao':list[9],
                    #'Lacre':list[10],
                    'afericao':format_value(list[11]),
                    'venda': 0
                })
            except Exception as e:
                print(f"Problema encontrado durante a leitura do arquivo {cnpj}.xlsx: {e}")
                continue

    list_of_dicts = []

    for i in range(end + 2, sheet.max_row):  # end + 2 to skip the header
        row_dict = {cell.column: cell.value for cell in sheet[i] if cell.value is not None}
        list_of_dicts.append(row_dict)

    tanque_data = []
    for list in list_of_dicts:
        if list:
            tanque_data.append({
                'type':'tanque',
                'tanque':list[1],
                #'Produto':list[2],
                'abertura':format_value(list[5]),
                'fechamento':format_value(list[8]),
                'recebimento':format_value(list[10]),
                'venda':0
            })

    bico_tanque_data = []
    for bico in bico_data:
        bico_tanque_data.append(bico)
    for tanque in tanque_data:
        bico_tanque_data.append(tanque)

    path_dac = f"input/dac/{cnpj}.txt"
    path_xlsx = f"output/{cnpj}.xlsx"
    return bico_tanque_data, cnpj, path_dac, path_xlsx