import openpyxl
from format_value import format_value

def get_data(cnpj, file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)

    sheet = wb.active

    init = None
    end = None

    for row in sheet.iter_rows():
        if "BOMBA" in str(row[0].value):
            init = row[0].row
        elif "TANTQUE" in str(row[0].value):
            end = row[0].row
            break  

    list_of_dicts = []

    if init is not None and end is not None and init < end:
        for i in range(init + 1, end):  # init + 1 to skip the header
            row_dict = {cell.column: cell.value for cell in sheet[i] if cell.value is not None}
            list_of_dicts.append(row_dict)  
    else:
        print("Não foi possível encontrar as linhas inicial ou final, ou o intervalo é inválido.")

    bico_data = []
    for list in list_of_dicts:
        if list:
            try:
                bico_data.append({
                    'type':'bico',
                    'bico':list[4],
                    #'Produto':list[5],
                    'abertura':list[6],
                    'fechamento':list[7],
                    #'Sem_intervencao':None,
                    #'Com_intervencao':None,
                    #'Lacre':list[3],
                    'afericao':list[9],
                    'venda':list[10]
                })
            except:
                pass
    
    list_of_dicts = []

    for i in range(end + 1, sheet.max_row):  # end + 1 to skip the header
        row_dict = {cell.column: cell.value for cell in sheet[i] if cell.value is not None}
        list_of_dicts.append(row_dict)

    tanque_data = []
    for list in list_of_dicts:
        if list:
            try:
                tanque_data.append({
                    'type':'tanque',
                    'tanque':list[1],
                    #'Produto':list[2],
                    'abertura':list[5],
                    'fechamento':list[7],
                    'recebimento':list[6],
                    'venda':0
                })
            except:
                pass

    bico_tanque_data = []
    for bico in bico_data:
        bico_tanque_data.append(bico)
    for tanque in tanque_data:
        bico_tanque_data.append(tanque)

    path_dac = f"input/dac/{cnpj}.txt"
    path_xlsx = f"output/{cnpj}.xlsx"
    return bico_tanque_data, cnpj, path_dac, path_xlsx