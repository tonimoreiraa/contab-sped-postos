import openpyxl

def save_to_sheet(bico_data, tanque_data, file_path):
    wb = openpyxl.Workbook()

    # Cria e preenche a planilha Bico
    ws_bico = wb.active
    ws_bico.title = 'Bico'

    # Define o cabeçalho (se necessário)
    headers = ["Bico", "Produto", "Abertura", "Fechamento", "Sem_intervencao", "Com_intervencao", "Afericao", "Obs_relatorio", "Obs_sped"]
    for col_num, header in enumerate(headers, start=1):
        ws_bico.cell(row=1, column=col_num, value=header)
    
    # Preenche as linhas com dados
    for row_num, data in enumerate(bico_data, start=2):
        for col_num, header in enumerate(headers, start=1):
            ws_bico.cell(row=row_num, column=col_num, value=data.get(header, ''))

    # Cria e preenche a planilha Tanque
    ws_tanque = wb.create_sheet(title='Tanque')

    # Define o cabeçalho (se necessário)
    headers_tanque = ["Tanque", "Produto", "Abertura", "Fechamento", "Recebimento", "Obs_relatorio", "Obs_sped"]
    for col_num, header in enumerate(headers_tanque, start=1):
        ws_tanque.cell(row=1, column=col_num, value=header)
    
    # Preenche as linhas com dados
    for row_num, data in enumerate(tanque_data, start=2):
        for col_num, header in enumerate(headers_tanque, start=1):
            ws_tanque.cell(row=row_num, column=col_num, value=data.get(header, ''))

    # Salva o arquivo Excel
    wb.save(file_path)

    print(f"Arquivo salvo em: {file_path}")
